from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from flask_cors import CORS
from datetime import datetime
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import os
import io
import sys
import logging
import traceback
from dotenv import load_dotenv
import openpyxl

# Load environment variables
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
    env_path = os.path.join(base_dir, '.env')
else:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env_path = os.path.join(base_dir, '.env')

load_dotenv(env_path)

# --- Session Secret Key Persistence ---
# To prevent logging users out when the application restarts
secret_key = os.getenv("SECRET_KEY")
if not secret_key:
    import secrets
    secret_key = secrets.token_hex(32)
    try:
        with open(env_path, "a") as env_file:
            env_file.write(f"\nSECRET_KEY={secret_key}\n")
    except Exception:
        pass

# --- Logging Setup ---
if getattr(sys, 'frozen', False):
    app_dir = os.path.dirname(sys.executable)
else:
    app_dir = os.path.dirname(os.path.abspath(__file__))

log_file = os.path.join(app_dir, 'debug.log')
logging.basicConfig(
    filename=log_file,
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logging.info("Starting Application (Enterprise Architecture)...")

# --- Path Helpers ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

if getattr(sys, 'frozen', False):
    template_dir = resource_path('templates')
    static_dir = resource_path('static')
else:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_dir = os.path.join(base_dir, '..', 'frontend', 'templates')
    static_dir = os.path.join(base_dir, '..', 'static')

import re

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
app.secret_key = secret_key

# Allow localhost and any local network subnet (192.168.x.x, 10.x.x.x, 172.16-31.x.x)
local_origin_re = re.compile(r"^https?://(localhost|127\.0\.0\.1|192\.168\.\d+\.\d+|10\.\d+\.\d+\.\d+|172\.(1[6-9]|2\d|3[0-1])\.\d+\.\d+)(:\d+)?$")
CORS(app, resources={r"/*": {"origins": [local_origin_re]}},
     supports_credentials=True)

# --- Database Service Setup ---
from db_service import DbService
MONGO_URI = os.getenv("MONGO_URI")
db_service = DbService(mongo_uri=MONGO_URI)

# --- Helpers ---
def current_user_id():
    """Return the logged-in user's ID string, or None."""
    return session.get('user_id')

def current_user_role():
    """Return the logged-in user's role, or None."""
    return session.get('role')

def is_admin():
    return current_user_role() == 'admin'

# --- Auth Decorators ---
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user_id():
            if request.is_json or request.path.startswith('/api'):
                return jsonify({"error": "Unauthorized", "redirect": "/login"}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user_id():
            return redirect(url_for('login_page'))
        if not is_admin():
            return jsonify({"error": "Forbidden — admin only"}), 403
        return f(*args, **kwargs)
    return decorated

# ─────────────────────────────────────────────
#  PAGE ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def home():
    if not current_user_id():
        return redirect(url_for('login_page'))
    return render_template('dashboard.html')

@app.route('/login')
def login_page():
    if current_user_id():
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route('/register')
def register_page():
    if current_user_id():
        return redirect(url_for('home'))
    return render_template('register.html')

@app.route('/admin')
@admin_required
def admin_page():
    return render_template('admin.html')

@app.route('/heartbeat', methods=['POST', 'GET'])
def heartbeat():
    # Attempt background sync periodically
    if not db_service.is_online:
        db_service.connect_mongodb()
    if db_service.is_online:
        try:
            db_service.sync_offline_data()
        except Exception as e:
            logging.error(f"Background heartbeat sync failed: {e}")
            
    return jsonify({
        "status": "ok",
        "database": "online" if db_service.is_online else "offline"
    })

# ─────────────────────────────────────────────
#  AUTH API ROUTES
# ─────────────────────────────────────────────

@app.route('/auth/register', methods=['POST'])
def auth_register():
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400

    username = (data.get('username') or '').strip()
    email    = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not username or not email or not password:
        return jsonify({"error": "All fields are required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if '@' not in email:
        return jsonify({"error": "Invalid email address"}), 400

    # First user ever → admin
    role = 'admin' if db_service.count_users() == 0 else 'user'

    try:
        user = db_service.create_user(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
            is_active=True
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 409

    session['user_id'] = user['id']
    session['username'] = username
    session['role'] = role

    return jsonify({
        "message": "Registered successfully",
        "role": role,
        "username": username
    }), 201

@app.route('/auth/login', methods=['POST'])
def auth_login():
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400

    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    user = db_service.get_user_by_username(username)
    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid username or password"}), 401

    if not user.get('is_active', True):
        return jsonify({"error": "Account is deactivated. Contact admin."}), 403

    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    logging.info(f"User logged in: {username}")
    
    return jsonify({
        "message": "Login successful",
        "username": user['username'],
        "role": user['role']
    })

@app.route('/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({"message": "Logged out"})

@app.route('/auth/me', methods=['GET'])
@login_required
def auth_me():
    return jsonify({
        "user_id": current_user_id(),
        "username": session.get('username'),
        "role": current_user_role()
    })

@app.route('/auth/forgot-password', methods=['POST'])
def auth_forgot_password():
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400

    username = (data.get('username') or '').strip()
    email    = (data.get('email') or '').strip().lower()
    new_password = data.get('new_password') or ''

    if not username or not email or not new_password:
        return jsonify({"error": "All fields are required"}), 400
    if len(new_password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400

    new_hash = generate_password_hash(new_password)
    success = db_service.reset_password(username, email, new_hash)
    if not success:
        return jsonify({"error": "Username and Email do not match our records"}), 400

    return jsonify({"message": "Password reset successfully"})

# ─────────────────────────────────────────────
#  ADMIN API ROUTES
# ─────────────────────────────────────────────

@app.route('/admin/stats', methods=['GET'])
@admin_required
def admin_stats():
    total_users = db_service.count_users()
    active_users = db_service.count_active_users()
    total_expenses = db_service.get_expenses_count()
    total_amount = db_service.get_expenses_total_amount()

    return jsonify({
        "total_users": total_users,
        "active_users": active_users,
        "total_expenses": total_expenses,
        "total_amount": total_amount
    })

@app.route('/admin/users', methods=['GET'])
@admin_required
def admin_list_users():
    users = db_service.list_users()
    result = []
    for u in users:
        exp_count = db_service.get_expenses_count(user_id=u['id'])
        result.append({
            "id": u['id'],
            "username": u.get('username'),
            "email": u.get('email'),
            "role": u.get('role', 'user'),
            "is_active": bool(u.get('is_active', True)),
            "created_at": u.get('created_at', ''),
            "expense_count": exp_count
        })
    return jsonify(result)

@app.route('/admin/users/<uid>', methods=['PUT'])
@admin_required
def admin_update_user(uid):
    if uid == current_user_id() and request.json.get('role') == 'user':
        return jsonify({"error": "You cannot remove your own admin role"}), 400

    data = request.json or {}
    update = {}
    if 'role' in data and data['role'] in ('admin', 'user'):
        update['role'] = data['role']
    if 'is_active' in data:
        update['is_active'] = bool(data['is_active'])

    if not update:
        return jsonify({"error": "Nothing to update"}), 400

    db_service.update_user(uid, update)
    return jsonify({"message": "User updated"})

@app.route('/admin/users/<uid>', methods=['DELETE'])
@admin_required
def admin_delete_user(uid):
    if uid == current_user_id():
        return jsonify({"error": "You cannot delete your own account"}), 400

    db_service.delete_user(uid)
    return jsonify({"message": "User and their data deleted"})

# ─────────────────────────────────────────────
#  DASHBOARD STATS (Strict Data Isolation)
# ─────────────────────────────────────────────

@app.route('/dashboard-stats', methods=['GET'])
@login_required
def dashboard_stats():
    uid = current_user_id()
    target_month = request.args.get('month')
    if not target_month:
        target_month = datetime.now().strftime("%b %Y")

    # Current isolated Salary
    current_salary = db_service.get_salary(uid, target_month)

    # Current isolated Expenses
    expenses, _ = db_service.query_expenses(user_id=uid, month=target_month, limit=99999)
    current_expenses = sum(e['amount'] for e in expenses)

    # Previous isolated Balance calculations
    salary_map = db_service.get_all_salaries_map(uid)
    expense_map = db_service.aggregate_expenses_by_month(uid)

    try:
        target_date_obj = datetime.strptime(target_month, "%b %Y")
    except ValueError:
        target_date_obj = datetime.now()

    previous_balance = 0.0
    all_months = set(list(salary_map.keys()) + list(expense_map.keys()))

    for m_str in all_months:
        if not m_str:
            continue
        try:
            m_date = datetime.strptime(m_str, "%b %Y")
            if m_date < target_date_obj:
                previous_balance += salary_map.get(m_str, 0.0) - expense_map.get(m_str, 0.0)
        except ValueError:
            logging.warning(f"Skipping malformed month string: '{m_str}'")
            continue

    total_available   = previous_balance + current_salary
    remaining_balance = total_available - current_expenses

    def _safe_month_key(m_str):
        try:
            return datetime.strptime(m_str, "%b %Y") if m_str else datetime.min
        except ValueError:
            return datetime.min

    available_months_set = sorted(list(all_months), key=_safe_month_key, reverse=True)

    # Automatically check and trigger background sync on loading stats
    if not db_service.is_online:
        db_service.connect_mongodb()
    if db_service.is_online:
        try:
            db_service.sync_offline_data()
        except Exception as e:
            logging.error(f"Stat-triggered sync error: {e}")

    return jsonify({
        "current_filter":   target_month,
        "salary":           current_salary,
        "previous_balance": previous_balance,
        "current_expenses": current_expenses,
        "total_available":  total_available,
        "remaining_balance": remaining_balance,
        "available_months": available_months_set,
        "database": "online" if db_service.is_online else "offline"
    })

@app.route('/stats/category', methods=['GET'])
@login_required
def category_stats():
    uid = current_user_id()
    month = request.args.get('month') or datetime.now().strftime("%b %Y")
    data = db_service.aggregate_expenses_by_category(uid, month)
    return jsonify(data)

# ─────────────────────────────────────────────
#  SALARY
# ─────────────────────────────────────────────

@app.route('/salary', methods=['POST'])
@login_required
def set_salary():
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400

    month  = data.get('month')
    amount = data.get('amount')
    if not month or amount is None:
        return jsonify({"error": "Missing data"}), 400

    try:
        amount_float = float(amount)
    except (ValueError, TypeError):
        return jsonify({"error": "amount must be a number"}), 400

    uid = current_user_id()
    res = db_service.set_salary(uid, month, amount_float)
    return jsonify(res)

# ─────────────────────────────────────────────
#  EXPENSES
# ─────────────────────────────────────────────

@app.route('/expenses', methods=['GET'])
@login_required
def get_expenses():
    uid  = current_user_id()
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    month_filter = request.args.get('month')
    search   = request.args.get('search')
    category = request.args.get('category')

    items, total = db_service.query_expenses(
        user_id=uid,
        month=month_filter,
        search=search,
        category=category,
        skip=(page - 1) * limit,
        limit=limit
    )
    has_next = (page * limit) < total

    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit,
        "has_next": has_next
    })

@app.route('/expenses', methods=['POST'])
@login_required
def add_expense():
    data = request.json
    if not data or 'item' not in data or 'amount' not in data:
        return jsonify({"error": "Invalid data"}), 400

    date_str  = data.get('date')
    month_str = data.get('month')

    if date_str:
        d = None
        for fmt in ["%Y-%m-%d", "%d %m %Y", "%d-%m-%Y", "%d/%m/%Y"]:
            try:
                d = datetime.strptime(date_str, fmt)
                date_str = d.strftime("%Y-%m-%d")
                break
            except ValueError:
                pass
        if d is None:
            return jsonify({"error": f"Unrecognised date format: '{date_str}'"}), 400
        if d:
            month_str = d.strftime("%b %Y")
            data['date'] = date_str

    if not month_str:
        month_str = datetime.now().strftime("%b %Y")

    # Safe float conversion
    try:
        amount = float(data['amount'])
    except (ValueError, TypeError):
        return jsonify({"error": "Amount must be a numeric value"}), 400

    uid = current_user_id()
    expense = db_service.add_expense(
        user_id=uid,
        item=data['item'],
        amount=amount,
        category=data.get('category', 'General'),
        date=date_str,
        month=month_str
    )
    return jsonify(expense), 201

@app.route('/expenses/<id>', methods=['PUT', 'DELETE'])
@login_required
def expense_op(id):
    uid = current_user_id()
    
    # Isolate: Verify item belongs to user before operating
    item = db_service.get_expense_by_id(id, uid)
    if not item:
        return jsonify({"error": "Not found or not authorised"}), 404

    if request.method == 'DELETE':
        db_service.delete_expense(id, uid)
        return jsonify({"message": "Deleted"})

    elif request.method == 'PUT':
        data = request.json
        if not data:
            return jsonify({"error": "Missing JSON body"}), 400

        update_fields = {}
        if 'item' in data:
            update_fields['item'] = data['item']
        if 'amount' in data:
            try:
                update_fields['amount'] = float(data['amount'])
            except (ValueError, TypeError):
                return jsonify({"error": "amount must be a number"}), 400
        if 'category' in data:
            update_fields['category'] = data['category']

        if 'date' in data:
            date_str = data['date']
            d = None
            for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"]:
                try:
                    d = datetime.strptime(date_str, fmt)
                    date_str = d.strftime("%Y-%m-%d")
                    break
                except ValueError:
                    pass
            if d:
                update_fields['date']  = date_str
                update_fields['month'] = d.strftime("%b %Y")
            else:
                return jsonify({"error": f"Unrecognised date format: '{date_str}'"}), 400

        db_service.update_expense(id, uid, update_fields)
        return jsonify(db_service.get_expense_by_id(id, uid))

@app.route('/expenses', methods=['DELETE'])
@login_required
def clear_expenses():
    try:
        uid = current_user_id()
        db_service.clear_all_user_expenses(uid)
        return jsonify({"message": "All data deleted"})
    except Exception as e:
        logging.error(f"clear_expenses error: {e}")
        return jsonify({"error": str(e)}), 500

# ─────────────────────────────────────────────
#  EXPORT & EXCEL IMPORT
# ─────────────────────────────────────────────

@app.route('/export', methods=['GET'])
@login_required
def export_excel():
    uid = current_user_id()
    month_filter = request.args.get('month')
    expenses = db_service.get_expenses_list_for_export(uid, month_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["ID", "Date", "Item", "Category", "Amount", "Month"])

    for e in expenses:
        ws.append([e.get('id'), e.get('date', ''), e.get('item'),
                   e.get('category', 'General'), e.get('amount'), e.get('month')])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Expenses_{month_filter}.xlsx" if month_filter else "All_Expenses.xlsx"
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/import', methods=['POST'])
@login_required
def import_expenses():
    if 'file' not in request.files:
        return jsonify({"error": "No file selected"}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "Please upload a valid Excel spreadsheet (.xlsx)"}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "Uploaded Excel spreadsheet is empty"}), 400

        headers = [str(h).strip().lower() for h in rows[0]]
        col_map = {}
        for col_idx, header in enumerate(headers):
            if 'item' in header:
                col_map['item'] = col_idx
            elif 'amount' in header:
                col_map['amount'] = col_idx
            elif 'category' in header:
                col_map['category'] = col_idx
            elif 'date' in header:
                col_map['date'] = col_idx

        required_cols = ['item', 'amount', 'date']
        missing = [c for c in required_cols if c not in col_map]
        if missing:
            return jsonify({"error": f"Missing required headers: {', '.join(missing)}"}), 400

        uid = current_user_id()
        imported_count = 0

        for row in rows[1:]:
            if not any(row):
                continue

            item = row[col_map['item']]
            amount_val = row[col_map['amount']]
            date_val = row[col_map['date']]
            category = row[col_map['category']] if 'category' in col_map else 'General'

            if not item or amount_val is None or not date_val:
                continue

            # Safe amount parsing
            try:
                amount = float(amount_val)
            except (ValueError, TypeError):
                continue

            # Safe category parsing
            category = str(category).strip() if category else 'General'
            if not category:
                category = 'General'

            # Parse date
            date_str = None
            d = None
            if isinstance(date_val, datetime):
                d = date_val
                date_str = d.strftime("%Y-%m-%d")
            else:
                date_val_str = str(date_val).strip()
                for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d %b %Y", "%b %Y"]:
                    try:
                        d = datetime.strptime(date_val_str, fmt)
                        date_str = d.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        pass

            if not d:
                continue

            month_str = d.strftime("%b %Y")
            db_service.add_expense(uid, str(item), amount, category, date_str, month_str)
            imported_count += 1

        return jsonify({"message": f"Successfully imported {imported_count} transaction(s)"})
    except Exception as e:
        logging.error(f"Excel import error: {e}")
        return jsonify({"error": f"Spreadsheet import failed: {str(e)}"}), 500

# ─────────────────────────────────────────────
#  CATEGORIES
# ─────────────────────────────────────────────

@app.route('/categories', methods=['GET'])
@login_required
def get_categories():
    cats = db_service.list_categories()
    return jsonify(cats)

@app.route('/categories', methods=['POST'])
@login_required
def add_category():
    data = request.json
    if not data:
        return jsonify({"error": "Missing JSON body"}), 400
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"error": "Missing name"}), 400
        
    try:
        cat = db_service.add_category(name)
        return jsonify(cat), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/categories/<id>', methods=['DELETE'])
@login_required
def delete_category(id):
    try:
        db_service.delete_category(id)
        return jsonify({"message": "Deleted"})
    except Exception as e:
        return jsonify({"error": str(e)}), 409

# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == '__main__':
    logging.info("Starting Server...")
    try:
        import threading

        def run_flask():
            app.run(debug=False, host='0.0.0.0', port=8000, use_reloader=False)

        flask_thread = threading.Thread(target=run_flask, daemon=True)
        flask_thread.start()

        from PyQt5.QtWidgets import QApplication, QMainWindow
        from PyQt5.QtWebEngineWidgets import QWebEngineView
        from PyQt5.QtCore import QUrl
        from PyQt5.QtGui import QIcon

        qt_app = QApplication(sys.argv)
        qt_app.setApplicationName("Padharia Expense Tracker")

        class MainWindow(QMainWindow):
            def __init__(self):
                super().__init__()
                self.setWindowTitle("Padharia Expense Tracker")
                self.resize(1280, 860)
                icon_path = resource_path(os.path.join('static', 'rupee.ico'))
                self.setWindowIcon(QIcon(icon_path))

                self.browser = QWebEngineView()
                self.browser.setUrl(QUrl("http://127.0.0.1:8000"))
                self.setCentralWidget(self.browser)

                self.browser.page().profile().downloadRequested.connect(self.on_download_requested)

            def on_download_requested(self, download):
                from PyQt5.QtWidgets import QFileDialog
                suggested_filename = download.suggestedFileName()
                path, _ = QFileDialog.getSaveFileName(
                    self, "Save File", suggested_filename, "Excel Files (*.xlsx);;All Files (*)")
                if path:
                    download.setPath(path)
                    download.accept()
                else:
                    download.cancel()

        window = MainWindow()
        window.show()
        logging.info("PyQt5 Window Launched")
        sys.exit(qt_app.exec_())

    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
        logging.critical(f"Server crash: {e}")
        logging.critical(traceback.format_exc())
        sys.exit(1)
